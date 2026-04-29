"""
Generate HMI discrete alarms for all devices from graph.json.
Output: HMIAlarms_All.xlsx

Аварії по типах механізмів (бітова маска FLTCode : WORD):
  BIT0  (1)    - FLT_BREAKER             - всі типи
  BIT1  (2)    - FLT_OVERFLOW            - Redler, Noria
  BIT2  (4)    - FLT_NO_RUNFB            - Redler, Noria (тахо)
  BIT3  (8)    - FLT_ALINGMENT           - Noria
  BIT4  (16)   - FLT_NO_FEEDBACK         - Redler, Noria, Fan, Separator, Feeder
  BIT5  (32)   - FLT_GATE_MOVE_TIMEOUT   - Gate2P
  BIT6  (64)   - FLT_GATE_POS_UNKNOWN    - Gate2P
  BIT7  (128)  - FLT_BOTH_SENSORS        - Gate2P
  BIT8  (256)  - FLT_STOP_TIMEOUT        - Redler, Noria, Fan, Separator, Feeder
  BIT9  (512)  - FLT_VALVE_MOVE_TIMEOUT  - Valve3P
  BIT10 (1024) - FLT_VALVE_POS_UNKNOWN   - Valve3P
  BIT11 (2048) - FLT_VALVE_MULTIPLE_POS  - Valve3P
"""

import json
import openpyxl

# ── Read graph.json ──
with open('graph.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

devices = data.get('devices', [])
print(f"Found {len(devices)} devices")

# ── Alarm definitions per device type ──
# Format: (bit, suffix, alarm_text)
ALARMS_BY_TYPE = {
    'Redler': [
        (0,  'Breaker',      "Спрацьовав АВ"),
        (1,  'Overflow',     "Датчик підпору"),
        (2,  'Speed',        "Датчик швидкості"),
        (4,  'Feedback',     "Відсутній зворотній зв'язок контактора"),
        (8,  'StopTimeout',  "Тайм-аут вибігу"),
    ],
    'Noria': [
        (0,  'Breaker',      "Спрацьовав АВ"),
        (1,  'Overflow',     "Датчик підпору"),
        (2,  'Speed',        "Датчик швидкості"),
        (3,  'Alingment',    "Датчик сходу стрічки"),
        (4,  'Feedback',     "Відсутній зворотній зв'язок контактора"),
        (8,  'StopTimeout',  "Тайм-аут вибігу"),
    ],
    'Fan': [
        (0,  'Breaker',      "Спрацьовав АВ"),
        (4,  'Feedback',     "Відсутній зворотній зв'язок контактора"),
        (8,  'StopTimeout',  "Тайм-аут вибігу"),
    ],
    'Separator': [
        (0,  'Breaker',      "Спрацьовав АВ"),
        (4,  'Feedback',     "Відсутній зворотній зв'язок контактора"),
        (8,  'StopTimeout',  "Тайм-аут вибігу"),
    ],
    'Feeder': [
        (0,  'Breaker',      "Спрацьовав АВ"),
        (4,  'Feedback',     "Відсутній зворотній зв'язок контактора"),
        (8,  'StopTimeout',  "Тайм-аут вибігу"),
    ],
    'Gate2P': [
        (0,  'Breaker',      "Спрацьовав АВ"),
        (5,  'MoveTimeout',  "Тайм-аут переміщення"),
        (6,  'PosUnknown',   "Невизначена позиція"),
        (7,  'BothSensors',  "Обидва кінцевики активні"),
    ],
    'Valve3P': [
        (0,  'Breaker',      "Спрацьовав АВ"),
        (9,  'MoveTimeout',  "Тайм-аут переміщення"),
        (10, 'PosUnknown',   "Невизначена позиція"),
        (11, 'MultiplePOS',  "Кілька датчиків позиції активні"),
    ],
    'Silos': [
        (13, 'HighLevel',  "Верхній рівень"),
    ],
    'Sushka': [
        (0, 'Breaker', "Спрацьовав АВ"),
    ],
    'ReceivingPit': [
        (0, 'Breaker', "Спрацьовав АВ"),
    ],
}

# ── Read template row from HMIAlarms.xlsx ──
tpl_wb = openpyxl.load_workbook('HMIAlarms.xlsx')
tpl_ws = tpl_wb.active
template_row = {}
for c in range(1, tpl_ws.max_column + 1):
    header = tpl_ws.cell(row=1, column=c).value
    template_row[header] = tpl_ws.cell(row=2, column=c).value
tpl_wb.close()

print(f"Template columns: {list(template_row.keys())}")

# ── Build alarm rows ──
alarm_rows = []
alarm_id = 1

# === Аварії механізмів ===
for dev in devices:
    dev_name = dev['name'].replace('.', '_')
    dev_type = dev.get('type', '')

    alarm_defs = ALARMS_BY_TYPE.get(dev_type)
    if not alarm_defs:
        print(f"  WARNING: unknown type '{dev_type}' for device '{dev_name}' — skipped")
        continue

    for bit, suffix, alarm_text in alarm_defs:
        row = {}
        for col_name, col_value in template_row.items():
            if col_name == 'ID':
                row[col_name] = alarm_id
            elif col_name == 'Name':
                row[col_name] = f"{dev_name}_{suffix}"
            elif col_name == 'Alarm text [en-US], Alarm text 1':
                row[col_name] = f"{alarm_text} {dev_name}"
            elif col_name == 'Trigger tag':
                row[col_name] = f"{dev_name}_FLTCode"
            elif col_name == 'Trigger bit':
                row[col_name] = bit
            else:
                row[col_name] = col_value
        alarm_rows.append(row)
        alarm_id += 1

# === Аварії маршрутів (5 маршрутів, бітова маска RS_ResultCode) ===
ROUTE_ALARMS = [
    (0,  'REJ_Safety',         'Маршрут відхилено — GlobalSafetyStop'),
    (1,  'REJ_Owner',          'Маршрут відхилено — механізм зайнятий'),
    (2,  'REJ_Contract',       'Маршрут відхилено — порушення контракту'),
    (3,  'REJ_NotReady',       'Маршрут відхилено — механізм не готовий'),
    (4,  'REJ_DuplicateStart', 'Маршрут відхилено — дублікат START'),
    (5,  'REJ_GrainNotSet',    'Маршрут відхилено — тип зерна не задано'),
    (6,  'REJ_GrainMismatch',  'Маршрут відхилено — неспівпадіння типу зерна'),
    (8,  'ABRT_Operator',      'Маршрут зупинено оператором'),
    (9,  'ABRT_Safety',        'Маршрут скасовано — GlobalSafetyStop'),
    (10, 'ABRT_Local',         'Маршрут скасовано — Local/Manual'),
    (11, 'ABRT_Fault',         'Маршрут скасовано — аварія механізму'),
    (12, 'ABRT_Owner',         'Маршрут скасовано — конфлікт власника'),
    (13, 'ABRT_StartFailed',   'Маршрут скасовано — помилка захоплення'),
]

for route_num in range(1, 6):
    tag = f"ResultCode_Route{route_num}"
    for bit, suffix, alarm_text in ROUTE_ALARMS:
        row = {}
        for col_name, col_value in template_row.items():
            if col_name == 'ID':
                row[col_name] = alarm_id
            elif col_name == 'Name':
                row[col_name] = f"Route{route_num}_{suffix}"
            elif col_name == 'Alarm text [en-US], Alarm text 1':
                row[col_name] = f"{alarm_text} (Маршрут {route_num})"
            elif col_name == 'Trigger tag':
                row[col_name] = tag
            elif col_name == 'Trigger bit':
                row[col_name] = bit
            else:
                row[col_name] = col_value
        alarm_rows.append(row)
        alarm_id += 1

print(f"Generated {len(alarm_rows)} alarm rows")

# ── Write to new workbook ──
headers = list(template_row.keys())

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "DiscreteAlarms"

for c, h in enumerate(headers, 1):
    ws.cell(row=1, column=c, value=h)

for r_idx, row_data in enumerate(alarm_rows, 2):
    for c_idx, h in enumerate(headers, 1):
        ws.cell(row=r_idx, column=c_idx, value=row_data.get(h))

output_file = 'HMIAlarms_All.xlsx'
wb.save(output_file)
print(f"Saved to {output_file}")

# ── Summary by type ──
print("\nAlarms per type:")
for t, defs in ALARMS_BY_TYPE.items():
    dev_cnt = sum(1 for d in devices if d.get('type') == t)
    if dev_cnt:
        print(f"  {t:15s}: {dev_cnt:3d} devices x {len(defs)} alarms = {dev_cnt * len(defs):4d} rows")
print(f"\nTotal: {len(alarm_rows)} alarms for {len(devices)} devices")
