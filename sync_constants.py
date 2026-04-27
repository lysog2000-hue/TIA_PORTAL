"""
sync_constants.py
Читает лист Constants из Constant_project.xlsx как источник истины,
сравнивает с Constant.csv и обновляет CSV.
"""
import openpyxl
import csv
import sys
sys.stdout.reconfigure(encoding='utf-8')

# =========================================================
# 1. Читаем Constant_project.xlsx лист Constants
# =========================================================
print("=== Constant_project.xlsx [Constants] ===")
wb = openpyxl.load_workbook('Constant_project.xlsx')
ws = wb['Constants']
cols = [str(c.value) for c in ws[1]]
print(f"Columns: {cols}")

project = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        d = dict(zip(cols, row))
        name = str(d.get('Name', '')).strip()
        project[name] = {
            'type':    str(d.get('Data Type', '')).strip(),
            'value':   str(d.get('Value', '')).strip(),
            'comment': str(d.get('Comment', '')).strip()
        }
wb.close()
print(f"Loaded: {len(project)} constants")

# =========================================================
# 2. Читаем Constant.csv
# =========================================================
print("\n=== Constant.csv ===")
with open('Constant.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f, delimiter=';')
    fieldnames = reader.fieldnames
    csv_rows = {r['Name'].strip(): r for r in reader}
print(f"Loaded: {len(csv_rows)} constants")

# =========================================================
# 3. Сравниваем
# =========================================================
print("\n=== DIFFERENCES ===")
diffs = []

for name, proj in project.items():
    if name in csv_rows:
        csv_val  = csv_rows[name].get('Value', '').strip()
        csv_type = csv_rows[name].get('Data Type', '').strip()
        if proj['value'] != csv_val:
            diffs.append(('VALUE', name, proj['type'], proj['value'], csv_type, csv_val))
        if proj['type'] != csv_type:
            diffs.append(('TYPE',  name, proj['type'], proj['value'], csv_type, csv_val))
    else:
        diffs.append(('MISSING', name, proj['type'], proj['value'], '', ''))

for name in csv_rows:
    if name not in project:
        diffs.append(('EXTRA', name, '', '', csv_rows[name].get('Data Type',''), csv_rows[name].get('Value','')))

if diffs:
    print(f"Found {len(diffs)} differences:")
    for d in diffs:
        if d[0] == 'VALUE':
            print(f"  VALUE  | {d[1]:40s} | project={d[3]:10s} | csv={d[5]}")
        elif d[0] == 'TYPE':
            print(f"  TYPE   | {d[1]:40s} | project={d[2]:10s} | csv={d[4]}")
        elif d[0] == 'MISSING':
            print(f"  +ADD   | {d[1]:40s} | {d[2]}={d[3]}")
        elif d[0] == 'EXTRA':
            print(f"  -EXTRA | {d[1]:40s} | csv only: {d[4]}={d[5]}")
else:
    print("No differences!")

# =========================================================
# 4. Обновляем Constant.csv
# =========================================================
print("\n=== Updating Constant.csv ===")
updated = 0
added = 0

for name, proj in project.items():
    if name in csv_rows:
        old_val  = csv_rows[name].get('Value', '').strip()
        old_type = csv_rows[name].get('Data Type', '').strip()
        if old_val != proj['value'] or old_type != proj['type']:
            csv_rows[name]['Value'] = proj['value']
            csv_rows[name]['Data Type'] = proj['type']
            updated += 1
            print(f"  Updated: {name} | {old_type}={old_val} → {proj['type']}={proj['value']}")
    else:
        new_row = {fn: '' for fn in fieldnames}
        new_row['Name']      = name
        new_row['Path']      = 'Constant'
        new_row['Data Type'] = proj['type']
        new_row['Value']     = proj['value']
        new_row['Comment']   = proj['comment'] if proj['comment'] != 'None' else ''
        csv_rows[name] = new_row
        added += 1
        print(f"  Added:   {name} = {proj['value']} ({proj['type']})")

with open('Constant.csv', 'w', encoding='utf-8', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=';')
    writer.writeheader()
    for row in csv_rows.values():
        writer.writerow(row)

print(f"\nResult: {updated} updated, {added} added")
print("Constant.csv saved.")
