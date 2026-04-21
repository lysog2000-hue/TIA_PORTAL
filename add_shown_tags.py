import openpyxl

# Load HMITags.xlsx
wb = openpyxl.load_workbook(r'c:\Users\lysog\Desktop\REpository\tia_repo\HMITags.xlsx')
ws = wb['Hmi Tags']

# Find the last row with data
last_row = ws.max_row
print(f'Current last row: {last_row}')

# Collect all existing Word tag rows (rows 2 to 187, 186 tags)
existing_tags = []
for row_idx in range(2, 188):  # 186 tags
    name = ws.cell(row=row_idx, column=1).value
    if name and name.endswith('_FLTCode'):
        # Collect all cell values for this row
        row_data = {}
        for col in range(1, 31):  # 30 columns
            row_data[col] = ws.cell(row=row_idx, column=col).value
        existing_tags.append((row_idx, name, row_data))

print(f'Found {len(existing_tags)} existing FLTCode tags')

# Add new _Shown Bool rows after existing data
new_row_start = 188
for idx, (orig_row, orig_name, orig_data) in enumerate(existing_tags):
    new_row = new_row_start + idx
    shown_name = orig_name.replace('_FLTCode', '_FLTCode_Shown')
    
    # Copy most fields from original row
    for col in range(1, 31):
        ws.cell(row=new_row, column=col).value = orig_data.get(col)
    
    # Modify specific fields for the _Shown tag
    ws.cell(row=new_row, column=1).value = shown_name  # Name
    ws.cell(row=new_row, column=5).value = 'Bool'      # DataType
    ws.cell(row=new_row, column=6).value = 'Bool'      # HMI DataType
    
    # Update PLC tag: change .FLTCode to .FLTCode_Shown
    plc_tag = orig_data.get(4, '')
    if plc_tag and isinstance(plc_tag, str) and '.FLTCode' in plc_tag:
        new_plc_tag = plc_tag.replace('.FLTCode', '.FLTCode_Shown')
        ws.cell(row=new_row, column=4).value = new_plc_tag
    
    print(f'Row {new_row}: {shown_name} -> {ws.cell(row=new_row, column=4).value}')

print(f'\nAdded {len(existing_tags)} new _Shown rows')

# Save
wb.save(r'c:\Users\lysog\Desktop\REpository\tia_repo\HMITags.xlsx')
print('File saved!')
