"""
check_full.py
Полная проверка соответствия проекта:
1. Константы vs SCL код (Force_Code биты, FLT коды)
2. Mechs.xlsx vs graph.json (количество механизмов)
3. HMITags.xlsx vs graph.json (теги статусов)
4. FLT_STOP_TIMEOUT значение в Constant.csv vs project_context
5. Новые константы Valve3P в SCL файлах
"""
import openpyxl, csv, json, os, sys, re
sys.stdout.reconfigure(encoding='utf-8')

ISSUES = []
def issue(cat, msg):
    ISSUES.append(f"[{cat}] {msg}")
    print(f"  !! [{cat}] {msg}")
def ok(msg):
    print(f"  OK: {msg}")

# =========================================================
# Загрузка данных
# =========================================================
with open('graph.json', 'r', encoding='utf-8') as f:
    graph = json.load(f)
devices = graph['devices']

constants = {}
with open('Constant.csv', 'r', encoding='utf-8') as f:
    for row in csv.DictReader(f, delimiter=';'):
        constants[row['Name'].strip()] = {'type': row['Data Type'].strip(), 'value': row['Value'].strip()}

# =========================================================
# 1. Проверка Force_Code битов в SCL файлах
# =========================================================
print("\n=== 1. Force_Code bits in SCL files ===")

EXPECTED_FORCE = {
    'FC_Redler.scl':    {'forceBreaker': 1, 'forceOverflow': 2, 'forceSpeed': 4, 'forceFeedback': 256, 'forceStopTimeout': 64},
    'FC_Noria.scl':     {'forceBreaker': 1, 'forceOverflow': 2, 'forceSpeed': 4, 'forceAlingment': 8, 'forceFeedback': 256, 'forceStopTimeout': 64},
    'FC_Fan.scl':       {'forceBreaker': 1, 'forceFeedback': 256, 'forceStopTimeout': 64},
    'FC_Separator.scl': {'forceBreaker': 1, 'forceFeedback': 256, 'forceStopTimeout': 64},
    'FC_Feeder.scl':    {'forceBreaker': 1, 'forceFeedback': 256, 'forceStopTimeout': 64},
}

for fname, expected in EXPECTED_FORCE.items():
    fpath = os.path.join('Mechs', fname)
    if not os.path.exists(fpath):
        issue('SCL', f"{fname} not found")
        continue
    with open(fpath, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()
    for var, mask in expected.items():
        pattern = rf'#{var}\s*:=\s*\(#B\.Force_Code\s+AND\s+(\d+)\)'
        m = re.search(pattern, content)
        if m:
            actual = int(m.group(1))
            if actual == mask:
                ok(f"{fname}: {var} = {mask}")
            else:
                issue('FORCE_BIT', f"{fname}: {var} expected AND {mask}, got AND {actual}")
        else:
            issue('FORCE_BIT', f"{fname}: {var} pattern not found")

# =========================================================
# 2. Проверка FLT кодов в SCL файлах
# =========================================================
print("\n=== 2. FLT codes in SCL files ===")

FLT_VALUES = {
    'FLT_NONE': '0', 'FLT_BREAKER': '11', 'FLT_OVERFLOW': '10',
    'FLT_NO_RUNFB': '12', 'FLT_NO_FEEDBACK': '14', 'FLT_ALINGMENT': '15',
    'FLT_STOP_TIMEOUT': '22', 'FLT_GATE_MOVE_TIMEOUT': '16', 'FLT_GATE_POS_UNKNOWN': '17'
}

for name, expected_val in FLT_VALUES.items():
    if name in constants:
        actual = constants[name]['value']
        if actual == expected_val:
            ok(f"{name} = {actual}")
        else:
            issue('FLT_VALUE', f"{name}: expected {expected_val}, got {actual}")
    else:
        issue('FLT_MISSING', f"{name} not in Constant.csv")

# =========================================================
# 3. Проверка STS_RUNNING = 2 в SCL файлах
# =========================================================
print("\n=== 3. STS_RUNNING usage in FC files ===")
scl_files = ['Mechs/FC_Redler.scl', 'Mechs/FC_Noria.scl', 'Mechs/FC_Fan.scl',
             'Mechs/FC_Separator.scl', 'Mechs/FC_Feeder.scl']

for fpath in scl_files:
    if not os.path.exists(fpath):
        continue
    with open(fpath, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()
    # Проверяем что используется STS_RUNNING а не магическое число 2
    if '"STS_RUNNING"' in content:
        ok(f"{os.path.basename(fpath)}: uses STS_RUNNING constant")
    else:
        issue('SCL_CONST', f"{os.path.basename(fpath)}: STS_RUNNING not found")

# =========================================================
# 4. Проверка Mechs.xlsx — константы COUNT
# =========================================================
print("\n=== 4. Mechs.xlsx COUNT constants vs graph.json ===")
wb = openpyxl.load_workbook('Mechs.xlsx')
ws = wb.active
mechs_consts = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        mechs_consts[str(row[0]).strip()] = row[3]  # Value column
wb.close()

# Считаем из graph.json
from collections import Counter
type_counts = Counter(d['type'] for d in devices)
type_to_const = {
    'Gate2P':      'GATES2P_COUNT',
    'Redler':      'REDLERS_COUNT',
    'Noria':       'NORIAS_COUNT',
    'Fan':         'FANS_COUNT',
    'Separator':   'SEPARATORS_COUNT',
    'Feeder':      'FEEDERS_COUNT',
    'Valve3P':     'VALVES3P_COUNT',
    'Silos':       'SILOS_COUNT',
    'Sushka':      'SUSHKAS_COUNT',
    'ReceivingPit':'RECEIVING_PITS_COUNT',
}

for dtype, const_name in type_to_const.items():
    graph_count = type_counts.get(dtype, 0)
    if const_name in mechs_consts:
        xlsx_val = mechs_consts[const_name]
        # COUNT = количество - 1 (верхняя граница массива)
        expected_val = graph_count - 1
        if xlsx_val == expected_val:
            ok(f"{const_name} = {xlsx_val} (graph: {graph_count} devices)")
        else:
            issue('COUNT', f"{const_name}: xlsx={xlsx_val}, graph={graph_count} devices (expected {expected_val})")
    else:
        issue('COUNT', f"{const_name} not in Mechs.xlsx")

# =========================================================
# 5. Проверка HMITags.xlsx — теги статусов механизмов
# =========================================================
print("\n=== 5. HMITags.xlsx — Status tags ===")
wb = openpyxl.load_workbook('HMITags.xlsx')
ws = wb.active
cols = [str(c.value) for c in ws[1]]
hmi_tags = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        d = dict(zip(cols, row))
        hmi_tags[str(d.get('Name', '')).strip()] = d
wb.close()
print(f"  HMI tags loaded: {len(hmi_tags)}")

# Проверяем теги статусов для RuntimeLogger
runtime_types = {'Redler', 'Noria', 'Fan', 'Separator', 'Feeder'}
runtime_devs = [d for d in devices if d['type'] in runtime_types]
missing_status = []
for dev in runtime_devs:
    name = dev['name'].replace('.', '_')
    tag = f"DB_Mechs_Mechs{{{dev['id']}}}_Status"
    # Ищем по PLC тегу
    plc_tag = f"DB_Mechs.Mechs[{dev['id']}].Status"
    found = any(str(t.get('PLC tag', '')) == plc_tag for t in hmi_tags.values())
    if not found:
        missing_status.append(f"{name} (slot {dev['id']})")

if missing_status:
    issue('HMI_STATUS', f"Missing Status tags ({len(missing_status)}): {missing_status[:5]}")
else:
    ok(f"All {len(runtime_devs)} runtime devices have Status tags in HMITags.xlsx")

# =========================================================
# 6. Проверка новых констант Valve3P в SCL
# =========================================================
print("\n=== 6. New Valve3P constants in FC_Valve3P.scl ===")
valve_path = 'Mechs/FC_Valve3P.scl'
if os.path.exists(valve_path):
    with open(valve_path, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()
    new_consts = ['CMD_VALVE_POS0', 'CMD_VALVE_POS1', 'CMD_VALVE_POS2',
                  'FLT_VALVE_MOVE_TIMEOUT', 'FLT_VALVE_POS_UNKNOWN']
    for c in new_consts:
        if f'"{c}"' in content:
            ok(f"FC_Valve3P uses {c}")
        else:
            issue('VALVE3P', f"FC_Valve3P does not use {c} (new constant added to CSV)")
else:
    issue('SCL', f"{valve_path} not found")

# =========================================================
# ИТОГ
# =========================================================
print("\n" + "="*60)
print(f"TOTAL ISSUES: {len(ISSUES)}")
if ISSUES:
    for i, iss in enumerate(ISSUES, 1):
        print(f"  {i}. {iss}")
else:
    print("All checks passed!")
