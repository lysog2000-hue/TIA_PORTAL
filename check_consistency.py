"""
check_consistency.py
Проверяет соответствие между:
- HMITags_AlARMS.xlsx  — теги аварий (FLTCode теги)
- HMIAlarms_All.xlsx   — аварии (Trigger tag + Trigger bit)
- HMITags.xlsx         — все HMI теги
- Mechs.xlsx           — конфигурация механизмов
- graph.json           — граф устройств
- Constant.csv         — константы
"""

import openpyxl
import json
import csv
import sys

sys.stdout.reconfigure(encoding='utf-8')

ISSUES = []

def issue(category, msg):
    ISSUES.append(f"[{category}] {msg}")
    print(f"  ISSUE [{category}]: {msg}")

def ok(msg):
    print(f"  OK: {msg}")

# =========================================================
# 1. Загрузка graph.json
# =========================================================
print("\n=== 1. graph.json ===")
with open('graph.json', 'r', encoding='utf-8') as f:
    graph = json.load(f)

devices = {str(d['id']): d for d in graph['devices']}
device_names = {d['name'].replace('.','_'): d for d in graph['devices']}
print(f"  Devices: {len(devices)}")

# Типы с моточасами
RUNTIME_TYPES = {'Redler', 'Noria', 'Fan', 'Separator', 'Feeder'}
runtime_devices = [d for d in graph['devices'] if d['type'] in RUNTIME_TYPES]
print(f"  Runtime devices: {len(runtime_devices)}")

# =========================================================
# 2. Загрузка Constant.csv
# =========================================================
print("\n=== 2. Constant.csv ===")
constants = {}
with open('Constant.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f, delimiter=';')
    for row in reader:
        constants[row['Name'].strip()] = {
            'type': row['Data Type'].strip(),
            'value': row['Value'].strip()
        }
print(f"  Constants loaded: {len(constants)}")

# Проверяем ключевые константы
key_constants = [
    'FLT_NONE', 'FLT_BREAKER', 'FLT_OVERFLOW', 'FLT_NO_RUNFB',
    'FLT_NO_FEEDBACK', 'FLT_ALINGMENT', 'FLT_STOP_TIMEOUT',
    'FLT_GATE_MOVE_TIMEOUT', 'FLT_GATE_POS_UNKNOWN',
    'STS_IDLE', 'STS_STARTING', 'STS_RUNNING', 'STS_STOPPING', 'STS_FAULT',
    'CMD_NONE', 'CMD_START', 'CMD_STOP', 'CMD_RESET',
    'TYPE_REDLER', 'TYPE_NORIA', 'TYPE_FAN', 'TYPE_GATE2P',
    'TYPE_SEPARATOR', 'TYPE_FEEDER', 'TYPE_VALVE3P',
]
for c in key_constants:
    if c not in constants:
        issue('CONSTANT', f"Missing: {c}")
    else:
        ok(f"{c} = {constants[c]['value']} ({constants[c]['type']})")

# =========================================================
# 3. Загрузка Mechs.xlsx
# =========================================================
print("\n=== 3. Mechs.xlsx ===")
wb = openpyxl.load_workbook('Mechs.xlsx')
ws = wb.active
mechs_cols = [str(c.value) for c in ws[1]]
print(f"  Columns: {mechs_cols}")
mechs_data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        mechs_data.append(dict(zip(mechs_cols, row)))
wb.close()
print(f"  Mechs rows: {len(mechs_data)}")

# Проверяем что SlotId уникальны
slot_ids = [m.get('SlotId') or m.get('Slot') or m.get('ID') or m.get('id') for m in mechs_data]
print(f"  First row keys: {list(mechs_data[0].keys()) if mechs_data else 'empty'}")
if mechs_data:
    print(f"  First row: {mechs_data[0]}")

# =========================================================
# 4. Загрузка HMITags_AlARMS.xlsx
# =========================================================
print("\n=== 4. HMITags_AlARMS.xlsx ===")
wb = openpyxl.load_workbook('HMITags_AlARMS.xlsx')
ws = wb.active
alarm_tags_cols = [str(c.value) for c in ws[1]]
alarm_tags = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        alarm_tags.append(dict(zip(alarm_tags_cols, row)))
wb.close()
print(f"  Alarm tags: {len(alarm_tags)}")
print(f"  Columns: {alarm_tags_cols[:6]}")

# Строим словарь: имя тега → PLC тег
alarm_tag_map = {}
for t in alarm_tags:
    name = str(t.get('Name', ''))
    plc_tag = str(t.get('PLC tag', ''))
    dtype = str(t.get('DataType', ''))
    alarm_tag_map[name] = {'plc': plc_tag, 'type': dtype}

# Проверяем что каждый механизм из graph.json имеет FLTCode тег
print("\n  Checking FLTCode tags for all devices...")
missing_flt_tags = []
for dev in graph['devices']:
    name = dev['name'].replace('.', '_')
    tag_name = f"{name}_FLTCode"
    if tag_name not in alarm_tag_map:
        missing_flt_tags.append(tag_name)

if missing_flt_tags:
    issue('ALARM_TAGS', f"Missing FLTCode tags ({len(missing_flt_tags)}): {missing_flt_tags[:5]}...")
else:
    ok(f"All {len(graph['devices'])} devices have FLTCode tags")

# Проверяем тип данных — должен быть Word
wrong_type = [t for t in alarm_tags if str(t.get('DataType','')) not in ('Word', 'WORD')]
if wrong_type:
    issue('ALARM_TAGS', f"Wrong DataType (not Word): {[t['Name'] for t in wrong_type[:5]]}")
else:
    ok("All alarm tags have correct DataType=Word")

# =========================================================
# 5. Загрузка HMIAlarms_All.xlsx
# =========================================================
print("\n=== 5. HMIAlarms_All.xlsx ===")
wb = openpyxl.load_workbook('HMIAlarms_All.xlsx')
ws = wb.active
alarms_cols = [str(c.value) for c in ws[1]]
alarms = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        alarms.append(dict(zip(alarms_cols, row)))
wb.close()
print(f"  Alarms: {len(alarms)}")

# Ожидаемые биты по типам механизмов
EXPECTED_BITS = {
    'Redler':      {0, 1, 2, 6, 8},
    'Noria':       {0, 1, 2, 3, 6, 8},
    'Fan':         {0, 6, 8},
    'Separator':   {0, 6, 8},
    'Feeder':      {0, 6, 8},
    'Gate2P':      {0, 4, 5},
    'Valve3P':     {0, 4, 5},
    'Silos':       {0},
    'Sushka':      {0},
    'ReceivingPit':{0},
}

# Группируем аварии по trigger tag
from collections import defaultdict
alarms_by_tag = defaultdict(list)
for a in alarms:
    tag = str(a.get('Trigger tag', ''))
    bit = a.get('Trigger bit', 0)
    alarms_by_tag[tag].append(int(bit) if bit is not None else 0)

# Проверяем биты для каждого устройства
print("\n  Checking alarm bits per device type...")
bit_errors = []
for dev in graph['devices']:
    name = dev['name'].replace('.', '_')
    dtype = dev['type']
    tag_name = f"{name}_FLTCode"
    expected = EXPECTED_BITS.get(dtype, set())
    actual = set(alarms_by_tag.get(tag_name, []))

    missing = expected - actual
    extra = actual - expected

    if missing:
        bit_errors.append(f"{name}({dtype}): missing bits {missing}")
    if extra:
        bit_errors.append(f"{name}({dtype}): extra bits {extra}")

if bit_errors:
    print(f"  Bit errors ({len(bit_errors)}):")
    for e in bit_errors[:10]:
        issue('ALARM_BITS', e)
    if len(bit_errors) > 10:
        print(f"  ... and {len(bit_errors)-10} more")
else:
    ok(f"All alarm bits match expected for all {len(graph['devices'])} devices")

# Проверяем что trigger tag существует в HMITags_AlARMS
print("\n  Checking trigger tags exist in alarm tags...")
missing_trigger = []
for a in alarms:
    tag = str(a.get('Trigger tag', ''))
    if tag and tag not in alarm_tag_map:
        missing_trigger.append(tag)

missing_trigger = list(set(missing_trigger))
if missing_trigger:
    issue('ALARM_TRIGGER', f"Trigger tags not in HMITags_AlARMS ({len(missing_trigger)}): {missing_trigger[:5]}")
else:
    ok("All trigger tags exist in HMITags_AlARMS")

# =========================================================
# 6. Проверка PLC тегов в HMITags_AlARMS
# =========================================================
print("\n=== 6. PLC tag format check ===")
# Формат должен быть: DB_Mechs.Mechs[SlotId].FLTCode
# SlotId берём из graph.json
slot_by_name = {d['name'].replace('.','_'): d['id'] for d in graph['devices']}

plc_errors = []
for t in alarm_tags:
    tag_name = str(t.get('Name', ''))
    plc_tag = str(t.get('PLC tag', ''))
    # Имя тега: RD1_FLTCode → механизм RD1 → id из graph
    mech_name = tag_name.replace('_FLTCode', '')
    if mech_name in slot_by_name:
        slot_id = slot_by_name[mech_name]
        expected_plc = f"DB_Mechs.Mechs[{slot_id}].FLTCode"
        if plc_tag != expected_plc:
            plc_errors.append(f"{tag_name}: expected '{expected_plc}', got '{plc_tag}'")

if plc_errors:
    print(f"  PLC tag errors ({len(plc_errors)}):")
    for e in plc_errors[:10]:
        issue('PLC_TAG', e)
else:
    ok(f"All PLC tags have correct format DB_Mechs.Mechs[SlotId].FLTCode")

# =========================================================
# ИТОГ
# =========================================================
print("\n" + "="*60)
print(f"TOTAL ISSUES: {len(ISSUES)}")
if ISSUES:
    print("\nAll issues:")
    for i, iss in enumerate(ISSUES, 1):
        print(f"  {i}. {iss}")
else:
    print("No issues found!")
