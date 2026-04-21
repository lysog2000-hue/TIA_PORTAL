"""
Generate alarm tags from graph.json — matching HMITags.xlsx format
- For each device: {name}_FLTCode (UInt)
- Dots in device names are replaced with underscores
"""

import json
import openpyxl

# Read graph.json
with open('graph.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Extract all device names, replace '.' with '_'
devices = data.get('devices', [])
device_names = [device['name'].replace('.', '_') for device in devices]

print(f"Found {len(device_names)} devices")

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Hmi Tags"

# Headers — exactly like HMITags.xlsx
headers = [
    "Name", "Path", "Connection", "PLC tag", "DataType",
    "HMI DataType", "Length", "Access Method", "Address", "Start value",
    "Persistency", "Substitute value", "ID tag", "Comment [en-US]",
    "Acquisition mode", "Acquisition cycle", "Limit Upper 2 Type",
    "Limit Upper 2", "Limit Lower 2 Type", "Limit Lower 2",
    "Linear scaling", "End value PLC", "Start value PLC", "End value HMI",
    "Start value HMI", "Gmp relevant", "Confirmation Type",
    "RequiredFunctionRights", "Mandatory Commenting", "Scope"
]

# Write headers
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)

# Default values — matching HMITags.xlsx exactly
nv = "<No Value>"  # No Value placeholder

row_num = 2

for device_name in device_names:
    tag_name = f"{device_name}_FLTCode"

    ws.cell(row=row_num, column=1, value=tag_name)             # Name
    ws.cell(row=row_num, column=2, value="Alarm_tag")           # Path
    ws.cell(row=row_num, column=3, value=nv)                    # Connection
    ws.cell(row=row_num, column=4, value=nv)                    # PLC tag
    ws.cell(row=row_num, column=5, value="UInt")                # DataType
    ws.cell(row=row_num, column=6, value="UInt")                # HMI DataType
    ws.cell(row=row_num, column=7, value=2)                     # Length
    ws.cell(row=row_num, column=8, value=nv)                    # Access Method
    ws.cell(row=row_num, column=9, value=nv)                    # Address
    ws.cell(row=row_num, column=10, value=nv)                   # Start value
    ws.cell(row=row_num, column=11, value="False")              # Persistency
    ws.cell(row=row_num, column=12, value=nv)                   # Substitute value
    ws.cell(row=row_num, column=13, value=0)                    # ID tag
    ws.cell(row=row_num, column=14, value=nv)                   # Comment [en-US]
    ws.cell(row=row_num, column=15, value="Cyclic in operation")# Acquisition mode
    ws.cell(row=row_num, column=16, value="T1s")                # Acquisition cycle
    ws.cell(row=row_num, column=17, value="None")               # Limit Upper 2 Type
    ws.cell(row=row_num, column=18, value=nv)                   # Limit Upper 2
    ws.cell(row=row_num, column=19, value="None")               # Limit Lower 2 Type
    ws.cell(row=row_num, column=20, value=nv)                   # Limit Lower 2
    ws.cell(row=row_num, column=21, value="False")              # Linear scaling
    ws.cell(row=row_num, column=22, value=10)                   # End value PLC
    ws.cell(row=row_num, column=23, value=0)                    # Start value PLC
    ws.cell(row=row_num, column=24, value=100)                  # End value HMI
    ws.cell(row=row_num, column=25, value=0)                    # Start value HMI
    ws.cell(row=row_num, column=26, value="False")              # Gmp relevant
    ws.cell(row=row_num, column=27, value="None")               # Confirmation Type
    ws.cell(row=row_num, column=28, value=nv)                   # RequiredFunctionRights
    ws.cell(row=row_num, column=29, value="False")              # Mandatory Commenting
    ws.cell(row=row_num, column=30, value="System-wide")        # Scope

    row_num += 1

# Save workbook
output_file = 'HMIAlarms.xlsx'
wb.save(output_file)

print(f"✓ Generated {len(device_names)} UInt tags for {len(device_names)} devices")
print(f"✓ Dots replaced with underscores in tag names")
print(f"✓ Saved to {output_file}")
